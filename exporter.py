"""
Exporta os resultados do MSL para um arquivo Excel profissional com:
  - Aba 1: Protocolo do MSL (questões, strings, critérios)
  - Aba 2: Todos os artigos coletados (antes da triagem)
  - Aba 3: Planilha de triagem (título/resumo)
  - Aba 4: Planilha de extração de dados (artigos aceitos)
  - Aba 5: Resumo / estatísticas
"""
from __future__ import annotations
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from searchers.base import Article


# ── Paleta de cores ──────────────────────────────────────────────────────────
COR_CABECALHO    = "1F4E79"   # azul escuro
COR_SUBTITULO    = "2E75B6"   # azul médio
COR_DESTAQUE     = "D6E4F0"   # azul muito claro
COR_ACEITO       = "C6EFCE"   # verde claro
COR_REJEITADO    = "FFC7CE"   # vermelho claro
COR_PENDENTE     = "FFEB9C"   # amarelo claro
BRANCO           = "FFFFFF"
CINZA_CLARO      = "F2F2F2"


def _cell_style(cell,
                bold=False,
                font_color=BRANCO,
                bg_color=None,
                size=11,
                wrap=False,
                h_align="left",
                v_align="center"):
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=size)
    if bg_color:
        cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(
        horizontal=h_align, vertical=v_align, wrap_text=wrap
    )


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _cabecalho(ws, row: int, cols: list[str], bg: str = COR_CABECALHO):
    for i, texto in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=texto)
        _cell_style(c, bold=True, bg_color=bg, h_align="center")
        c.border = _thin_border()


def _auto_largura(ws, min_w=12, max_w=60):
    for col in ws.columns:
        melhor = min_w
        for cell in col:
            if cell.value:
                melhor = min(max(melhor, len(str(cell.value)) + 2), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = melhor


# ── Aba 1: Protocolo ─────────────────────────────────────────────────────────

def _aba_protocolo(wb, config_dict: dict):
    ws = wb.create_sheet("  Protocolo MSL")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    linhas = [
        ("PROTOCOLO DO MAPEAMENTO SISTEMÁTICO DA LITERATURA", None, True, COR_CABECALHO, 14),
        ("", None, False, None, 11),
        ("TÍTULO DA DISSERTAÇÃO", config_dict["DISSERTATION"]["titulo"], True, COR_SUBTITULO, 11),
        ("OBJETIVO GERAL", config_dict["DISSERTATION"]["objetivo_geral"], False, COR_DESTAQUE, 11),
        ("HIPÓTESE", config_dict["DISSERTATION"]["hipotese"], False, COR_DESTAQUE, 11),
        ("", None, False, None, 11),
        ("QUESTÕES DE PESQUISA", "", True, COR_SUBTITULO, 12),
    ]

    r = 1
    for label, valor, bold, bg, sz in linhas:
        ca = ws.cell(row=r, column=1, value=label)
        _cell_style(ca, bold=bold, bg_color=bg or BRANCO,
                    font_color=BRANCO if bg and bg not in (COR_DESTAQUE, BRANCO) else "000000",
                    size=sz, wrap=True)
        if valor is not None:
            cb = ws.cell(row=r, column=2, value=valor)
            _cell_style(cb, bg_color=bg or BRANCO,
                        font_color=BRANCO if bg and bg not in (COR_DESTAQUE, BRANCO) else "000000",
                        size=sz, wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2) if not valor else None
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.row_dimensions[r].height = 28 if sz >= 12 else 48
        r += 1

    for rq in config_dict["RESEARCH_QUESTIONS"]:
        ca = ws.cell(row=r, column=1, value=rq["id"])
        _cell_style(ca, bold=True, bg_color=COR_DESTAQUE, font_color="000000")
        cb = ws.cell(row=r, column=2, value=rq["questao"])
        _cell_style(cb, bg_color=CINZA_CLARO, font_color="000000", wrap=True)
        ws.row_dimensions[r].height = 36
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="STRINGS DE BUSCA").font = Font(bold=True, size=12, color=BRANCO)
    ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=COR_SUBTITULO)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    for ss in config_dict["SEARCH_STRINGS"]:
        ca = ws.cell(row=r, column=1, value=f"{ss['id']} — {ss['descricao']}")
        _cell_style(ca, bold=True, bg_color=COR_DESTAQUE, font_color="000000", wrap=True)
        cb = ws.cell(row=r, column=2, value=ss["string"])
        _cell_style(cb, font_color="000000", wrap=True)
        ws.row_dimensions[r].height = 48
        r += 1

    r += 1
    for titulo_secao, criterios in [
        ("CRITÉRIOS DE INCLUSÃO", config_dict["INCLUSION_CRITERIA"]),
        ("CRITÉRIOS DE EXCLUSÃO", config_dict["EXCLUSION_CRITERIA"]),
        ("CRITÉRIOS DE QUALIDADE", config_dict["QUALITY_CRITERIA"]),
    ]:
        ws.cell(row=r, column=1, value=titulo_secao).font = Font(bold=True, size=12, color=BRANCO)
        ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=COR_SUBTITULO)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
        for crit in criterios:
            id_c, desc = crit.split(": ", 1)
            ws.cell(row=r, column=1, value=id_c).font = Font(bold=True, color="000000")
            ws.cell(row=r, column=2, value=desc).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1


# ── Aba 2: Artigos coletados ──────────────────────────────────────────────────

def _aba_artigos(wb, artigos: list[Article], titulo_aba: str, incluir_triagem: bool = False):
    ws = wb.create_sheet(titulo_aba)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    colunas_base = [
        "Título", "Autores", "Ano", "Fonte (Base)", "Veículo",
        "Tipo", "String de Busca", "DOI", "URL", "Resumo",
        "Palavras-chave", "Citações",
    ]
    colunas_triagem = [
        "Triagem Título/Resumo", "Justificativa Triagem",
        "Triagem Texto Completo", "Justificativa Texto Completo",
        "Critério de Exclusão",
        "QC1 - Formalização", "QC2 - Comparação", "QC3 - Métricas", "QC4 - Reprodutibilidade",
        "Notas",
    ]

    cols = colunas_base + (colunas_triagem if incluir_triagem else [])
    _cabecalho(ws, 1, cols)

    larguras = {
        "Título": 50, "Autores": 35, "Ano": 8, "Fonte (Base)": 18, "Veículo": 30,
        "Tipo": 18, "String de Busca": 14, "DOI": 30, "URL": 35, "Resumo": 70,
        "Palavras-chave": 35, "Citações": 10,
        "Triagem Título/Resumo": 20, "Justificativa Triagem": 35,
        "Triagem Texto Completo": 22, "Justificativa Texto Completo": 35,
        "Critério de Exclusão": 20,
        "QC1 - Formalização": 18, "QC2 - Comparação": 18,
        "QC3 - Métricas": 18, "QC4 - Reprodutibilidade": 20, "Notas": 30,
    }
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 20)

    for r_idx, art in enumerate(artigos, start=2):
        d = art.to_dict()
        bg = CINZA_CLARO if r_idx % 2 == 0 else BRANCO

        for c_idx, col in enumerate(cols, start=1):
            val = d.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            fc = "000000"
            cell_bg = bg

            if incluir_triagem and col == "Triagem Título/Resumo":
                if val == "Aceito":
                    cell_bg = COR_ACEITO
                elif val == "Rejeitado":
                    cell_bg = COR_REJEITADO
                elif val == "Pendente":
                    cell_bg = COR_PENDENTE

            _cell_style(cell, bg_color=cell_bg, font_color=fc, wrap=(col in ("Resumo", "Notas")))
            cell.border = _thin_border()

        ws.row_dimensions[r_idx].height = 18

    # Habilitar AutoFilter
    ws.auto_filter.ref = ws.dimensions


# ── Aba 5: Resumo ─────────────────────────────────────────────────────────────

def _aba_resumo(wb, artigos_brutos: list[Article], artigos_dedup: list[Article],
                contagem_por_base: dict, contagem_por_string: dict):
    ws = wb.create_sheet("  Resumo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    def linha(r, label, valor, bold_val=False, bg=None):
        ca = ws.cell(row=r, column=1, value=label)
        cb = ws.cell(row=r, column=2, value=valor)
        _cell_style(ca, bold=True, bg_color=bg or CINZA_CLARO, font_color="000000")
        _cell_style(cb, bold=bold_val, bg_color=bg or BRANCO, font_color="000000", h_align="center")
        ws.row_dimensions[r].height = 20
        return r + 1

    r = 1
    titulo = ws.cell(row=r, column=1, value="RESUMO DA BUSCA — MSL")
    _cell_style(titulo, bold=True, bg_color=COR_CABECALHO, size=13)
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 30
    r += 2

    r = linha(r, "Data de execução", datetime.now().strftime("%d/%m/%Y %H:%M"))
    r = linha(r, "Total bruto (com duplicatas)", len(artigos_brutos), bold_val=True, bg=COR_DESTAQUE)
    r = linha(r, "Total após deduplicação", len(artigos_dedup), bold_val=True, bg=COR_DESTAQUE)
    r = linha(r, "Duplicatas removidas", len(artigos_brutos) - len(artigos_dedup))
    r += 1

    ws.cell(row=r, column=1, value="RESULTADOS POR BASE DE DADOS").font = Font(bold=True, color=BRANCO, size=11)
    ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=COR_SUBTITULO)
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    for base, qtd in sorted(contagem_por_base.items(), key=lambda x: -x[1]):
        r = linha(r, base, qtd)

    r += 1
    ws.cell(row=r, column=1, value="RESULTADOS POR STRING DE BUSCA").font = Font(bold=True, color=BRANCO, size=11)
    ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=COR_SUBTITULO)
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    for ss_id, qtd in sorted(contagem_por_string.items()):
        r = linha(r, ss_id, qtd)


# ── Função principal ──────────────────────────────────────────────────────────

def exportar_excel(
    artigos_brutos: list[Article],
    artigos_dedup: list[Article],
    config_dict: dict,
    caminho_saida: str,
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove aba padrão vazia

    # Contagens para resumo
    contagem_base: dict[str, int] = {}
    contagem_string: dict[str, int] = {}
    for art in artigos_dedup:
        contagem_base[art.fonte] = contagem_base.get(art.fonte, 0) + 1
        contagem_string[art.string_busca_id] = contagem_string.get(art.string_busca_id, 0) + 1

    _aba_protocolo(wb, config_dict)
    _aba_artigos(wb, artigos_brutos, "  Busca Bruta")
    _aba_artigos(wb, artigos_dedup, "  Sem Duplicatas", incluir_triagem=True)
    _aba_resumo(wb, artigos_brutos, artigos_dedup, contagem_base, contagem_string)

    wb.save(caminho_saida)
    print(f"\n    Excel salvo em: {caminho_saida}")
